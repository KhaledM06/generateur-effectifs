from flask import Flask, request, send_file, render_template_string, jsonify
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import io, os

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# ── Config ────────────────────────────────────────────────────────────────────
# Mots-clés pour détecter les protéines du salade bar
# Indépendant des quantités (7CVT, 10CVT, etc.)
SB_PROTEIN_KEYWORDS = [
    'POULET', 'VOLAILLE', 'DINDE', 'CANARD',
    'THON', 'SAUMON', 'CREVETTE', 'CABILLAUD', 'MAQUEREAU',
    'OEUF',
    'MOZZARELLA', 'FETA', 'CHEVRE', 'FROMAGE',
    'JAMBON', 'MORTADELLE', 'CHORIZO', 'BACON', 'LARDONS',
    'TOFU', 'FALAFEL', 'CHILI', 'EDAMAME', 'NEM',
    'BOEUF', 'VEAU', 'AGNEAU',
]
# Exclusions explicites (faux positifs dus aux mots-clés)
SB_PROTEIN_EXCLUSIONS = {
    'CIBOULETTE',  # contient "BOULET" mais n'est pas une proteine
}

def is_sb_protein(recette):
    """Détecte si une recette salade bar est une protéine par mots-clés."""
    rec_up = recette.upper().strip()
    if rec_up in SB_PROTEIN_EXCLUSIONS:
        return False
    return any(mot in rec_up for mot in SB_PROTEIN_KEYWORDS)
CATEGORY_ORDER = [
    'ENTREES','SALADE BAR','GRANDES SALADES','GRAND SANDWICH','SANDWICH CHAUD','MINI SANDWICH',
    'PLAT VIANDE/VOLAILLE 🍗🥩','PLAT VIANDE/VOLAILLE - CONSIGNE 🍗🥩','PLAT VIANDE/VOLAILLE - YSL 🍗🥩','PLAT VIANDE/VOLAILLE - CUISSON MINUTE 🍗🥩',
    'PLAT POISSON 🐟','PLAT POISSON - CONSIGNE 🐟','PLAT POISSON - YSL 🐟','PLAT POISSON - CUISSON MINUTE 🐟',
    'PLAT VEGGIE 🥦','PLAT VEGGIE - CONSIGNE 🥦','PLAT VEGGIE - YSL 🥦',
    'GARNITURE LEGUME','GARNITURE FECULENT','SOUPE','TARTES',
    'HEALTHY CUP','DESSERT','PATISSERIE','SALADE DE FRUIT','JUS DE FRUITS',
]
CANON = {
    'PLAT VIANDE/VOLAILLE 🍗🥩':'PLAT VIANDE','PLAT VIANDE/VOLAILLE - CONSIGNE 🍗🥩':'PLAT VIANDE',
    'PLAT VIANDE/VOLAILLE - YSL 🍗🥩':'PLAT VIANDE','PLAT VIANDE/VOLAILLE - CUISSON MINUTE 🍗🥩':'PLAT VIANDE',
    'PLAT POISSON 🐟':'PLAT POISSON','PLAT POISSON - CONSIGNE 🐟':'PLAT POISSON',
    'PLAT POISSON - YSL 🐟':'PLAT POISSON','PLAT POISSON - CUISSON MINUTE 🐟':'PLAT POISSON',
    'PLAT VEGGIE 🥦':'PLAT VEGGIE','PLAT VEGGIE - CONSIGNE 🥦':'PLAT VEGGIE','PLAT VEGGIE - YSL 🥦':'PLAT VEGGIE',
    'JUS DE FRUITS':'JUS','TARTES':'TARTES',
    'ENTREES':'ENTRÉES','GRANDES SALADES':'GRANDES SALADES','GRAND SANDWICH':'GRAND SANDWICH',
    'SANDWICH CHAUD':'SANDWICH CHAUD','MINI SANDWICH':'MINI SANDWICH','SALADE BAR':'SALADE BAR',
    'GARNITURE LEGUME':'GARNITURE LÉGUME','GARNITURE FECULENT':'GARNITURE FÉCULENT',
    'SOUPE':'SOUPE','HEALTHY CUP':'HEALTHY CUP','DESSERT':'DESSERT',
    'PATISSERIE':'PÂTISSERIE','SALADE DE FRUIT':'SALADE DE FRUIT',
}
CANON_ORDER = []
_seen = set()
for e in CATEGORY_ORDER:
    c = CANON.get(e, e)
    if c not in _seen:
        CANON_ORDER.append(c)
        _seen.add(c)

CAT_STYLE = {
    'ENTRÉES':           ('DBEAFE','1D4ED8'),
    'SALADE BAR':        ('DCFCE7','15803D'),'GRANDES SALADES':   ('DCFCE7','15803D'),
    'GRAND SANDWICH':    ('FEF3C7','92400E'),'SANDWICH CHAUD':    ('FEF3C7','92400E'),'MINI SANDWICH':     ('FEF3C7','92400E'),
    'PLAT VIANDE':       ('FEE2E2','991B1B'),'PLAT POISSON':      ('FEE2E2','991B1B'),'PLAT VEGGIE':       ('FEE2E2','991B1B'),
    'GARNITURE LÉGUME':  ('F3F4F6','374151'),'GARNITURE FÉCULENT':('F3F4F6','374151'),
    'SOUPE':             ('EDE9FE','5B21B6'),'TARTES':            ('EDE9FE','5B21B6'),
    'HEALTHY CUP':       ('FCE7F3','9D174D'),'DESSERT':           ('FCE7F3','9D174D'),'PÂTISSERIE':        ('FCE7F3','9D174D'),
    'SALADE DE FRUIT':   ('D1FAE5','065F46'),'JUS':               ('D1FAE5','065F46'),
}
POST_VENTE = ['Commandé','Reçu','Vendu','Reste cuit','Reste non cuit','Reste revendu','Pertes']

def S(style, color='E2E8F0'):
    return Side(style=style, color=color)

def generer_excel(fichier_source):
    """Génère le fichier Excel SharePoint formaté à partir de l'extraction."""
    df = pd.read_excel(fichier_source, skiprows=5)
    df = df.dropna(subset=['Recette'])
    df['Date du menu'] = df['Date du menu'].astype(str)
    df['Quantité réelle'] = pd.to_numeric(df['Quantité réelle'], errors='coerce').fillna(0).astype(int)
    df = df[df['Élément de repas'].isin(CATEGORY_ORDER)]
    mask_sb = df['Élément de repas'] == 'SALADE BAR'
    df = df[~mask_sb | (mask_sb & df['Recette'].apply(is_sb_protein))]
    df['canon'] = df['Élément de repas'].map(lambda x: CANON.get(x, x))

    dates = sorted(df['Date du menu'].unique(), key=lambda x: x.split('/')[::-1])
    sites = sorted(df['Site'].unique())
    n_sites = len(sites)
    N = len(POST_VENTE)

    data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))
    elems_per_date = defaultdict(list)
    for _, row in df.iterrows():
        d, c, r, s, q = row['Date du menu'], row['canon'], row['Recette'].strip(), row['Site'], row['Quantité réelle']
        data[d][c][r][s] += q
        if c not in elems_per_date[d]:
            elems_per_date[d].append(c)
    for d in dates:
        elems_per_date[d].sort(key=lambda e: CANON_ORDER.index(e) if e in CANON_ORDER else 999)

    def site_col(si, fi):
        return 2 + si * N + fi

    last_col = site_col(n_sites - 1, N - 1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Effectifs post-vente"
    ws.sheet_view.showGridLines = False

    d_start, d_end = dates[0], dates[-1]

    # Ligne 1 : Titre
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, "EFFECTIFS PAR PLAT — QUALITÉ & CO LABO")
    c.font = Font(name='Arial', bold=True, size=13, color='0F172A')
    c.fill = PatternFill('solid', fgColor='F8FAFC')
    c.alignment = Alignment(horizontal='left', vertical='center')

    # Ligne 2 : Sous-titre
    ws.row_dimensions[2].height = 16
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c2 = ws.cell(2, 1, f"{d_start} – {d_end}  •  Salade bar : protéines uniquement  •  Pertes : saisie manuelle")
    c2.font = Font(name='Arial', size=9, color='64748B')
    c2.fill = PatternFill('solid', fgColor='F8FAFC')
    c2.alignment = Alignment(horizontal='left', vertical='center')

    # Ligne 3 : Noms restaurants
    ws.row_dimensions[3].height = 20
    ws.cell(3, 1).fill = PatternFill('solid', fgColor='1E293B')
    for si, site in enumerate(sites):
        sc = site_col(si, 0)
        ec = site_col(si, N - 1)
        ws.merge_cells(start_row=3, start_column=sc, end_row=3, end_column=ec)
        c = ws.cell(3, sc, site.replace('RE ', ''))
        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1E293B')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = Border(left=S('medium','0F172A'), right=S('medium','0F172A'))

    # Ligne 4 : En-têtes colonnes
    ws.row_dimensions[4].height = 28
    ws.cell(4, 1).fill = PatternFill('solid', fgColor='334155')
    ws.cell(4, 1).border = Border(bottom=S('medium','94A3B8'))
    for si in range(n_sites):
        for fi, fname in enumerate(POST_VENTE):
            col = site_col(si, fi)
            is_perte = (fi == N - 1)
            c = ws.cell(4, col, fname)
            c.font = Font(name='Arial', bold=True, size=8,
                         color='FDE68A' if is_perte else 'F1F5F9')
            c.fill = PatternFill('solid', fgColor='334155')
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = Border(
                left=S('medium','475569') if fi == 0 else S('hair','475569'),
                right=S('medium','475569') if is_perte else S('hair','475569'),
                bottom=S('medium','94A3B8')
            )

    ws.freeze_panes = 'B5'

    current_row = 5
    for date in dates:
        # Séparateur de date
        ws.row_dimensions[current_row].height = 18
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col)
        cd = ws.cell(current_row, 1, f"  {date}")
        cd.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cd.fill = PatternFill('solid', fgColor='0F172A')
        cd.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        for elem in elems_per_date[date]:
            bg, fg = CAT_STYLE.get(elem, ('F3F4F6', '374151'))
            cat_fill = PatternFill('solid', fgColor=bg)
            cat_font = Font(name='Arial', bold=True, size=9, color=fg)
            rec_font = Font(name='Arial', size=9, color='374151')

            # Ligne catégorie
            ws.row_dimensions[current_row].height = 16
            cat_by_site = defaultdict(int)
            for rec in data[date][elem]:
                for s, q in data[date][elem][rec].items():
                    cat_by_site[s] += q

            ca = ws.cell(current_row, 1, f"  {elem}")
            ca.font = cat_font; ca.fill = cat_fill
            ca.alignment = Alignment(horizontal='left', vertical='center')
            ca.border = Border(top=S('thin','CBD5E1'), bottom=S('thin','CBD5E1'))

            for si, site in enumerate(sites):
                col = site_col(si, 0)
                val = cat_by_site.get(site) or None
                cv = ws.cell(current_row, col, val)
                cv.font = cat_font; cv.fill = cat_fill
                cv.alignment = Alignment(horizontal='right', vertical='center')
                cv.border = Border(left=S('medium','94A3B8'), top=S('thin','CBD5E1'), bottom=S('thin','CBD5E1'))
                for fi in range(1, N):
                    cell = ws.cell(current_row, col + fi)
                    cell.fill = cat_fill
                    cell.border = Border(
                        left=S('hair'), top=S('thin','CBD5E1'), bottom=S('thin','CBD5E1'),
                        right=S('medium','94A3B8') if fi == N-1 else S('hair')
                    )
            current_row += 1

            # Lignes recettes
            for rec in sorted(data[date][elem].keys()):
                ws.row_dimensions[current_row].height = 15
                ra = ws.cell(current_row, 1, f"    {rec}")
                ra.font = rec_font
                ra.fill = PatternFill('solid', fgColor='FFFFFF')
                ra.alignment = Alignment(horizontal='left', vertical='center')
                ra.border = Border(top=S('hair','F1F5F9'), bottom=S('hair','F1F5F9'))

                for si, site in enumerate(sites):
                    col = site_col(si, 0)
                    qty = data[date][elem][rec].get(site) or None
                    cv = ws.cell(current_row, col, qty)
                    cv.font = Font(name='Arial', size=9, color='374151')
                    cv.fill = PatternFill('solid', fgColor='FFFFFF')
                    cv.alignment = Alignment(horizontal='right', vertical='center')
                    cv.border = Border(left=S('medium','94A3B8'), top=S('hair','F1F5F9'), bottom=S('hair','F1F5F9'))
                    cv.number_format = '#,##0'
                    for fi in range(1, N - 1):
                        cell = ws.cell(current_row, col + fi)
                        cell.fill = PatternFill('solid', fgColor='FFFFFF')
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.border = Border(left=S('hair'), top=S('hair','F1F5F9'), bottom=S('hair','F1F5F9'))
                        cell.number_format = '#,##0'
                    pc = ws.cell(current_row, col + N - 1)
                    pc.font = Font(name='Arial', size=9, color='92400E')
                    pc.fill = PatternFill('solid', fgColor='FFFBF5')
                    pc.alignment = Alignment(horizontal='right', vertical='center')
                    pc.border = Border(left=S('hair'), right=S('medium','94A3B8'),
                                      top=S('hair','F1F5F9'), bottom=S('hair','F1F5F9'))
                    pc.number_format = '#,##0'
                current_row += 1

        ws.row_dimensions[current_row].height = 6
        for col in range(1, last_col + 1):
            ws.cell(current_row, col).fill = PatternFill('solid', fgColor='F8FAFC')
        current_row += 1

    # Largeurs
    ws.column_dimensions['A'].width = 70
    pv_widths = [9, 5, 7, 8, 9, 10, 6]
    for si in range(n_sites):
        for fi, w in enumerate(pv_widths):
            ws.column_dimensions[get_column_letter(site_col(si, fi))].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nom_fichier = f"Effectifs_SharePoint_{d_start.replace('/','')}_au_{d_end.replace('/','')}.xlsx"
    return output, nom_fichier


HTML = '''<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Générateur Effectifs — Qualité & Co Labo</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;background:#F8FAFC;color:#1E293B;min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:24px}
.card{background:#fff;border:1px solid #E2E8F0;border-radius:16px;padding:40px;max-width:540px;width:100%;box-shadow:0 4px 24px rgba(0,0,0,.06)}
.logo{background:#0F172A;border-radius:10px;padding:10px 16px;display:inline-block;margin-bottom:24px}
.logo span{color:#fff;font-size:13px;font-weight:600;letter-spacing:.5px}
h1{font-size:22px;font-weight:600;color:#0F172A;margin-bottom:6px}
.subtitle{font-size:13px;color:#64748B;margin-bottom:32px;line-height:1.5}
.drop-zone{border:2px dashed #CBD5E1;border-radius:12px;padding:40px 24px;text-align:center;cursor:pointer;transition:all .2s;margin-bottom:20px;position:relative}
.drop-zone:hover,.drop-zone.over{border-color:#3B82F6;background:#EFF6FF}
.drop-zone input{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.drop-icon{font-size:36px;margin-bottom:12px}
.drop-text{font-size:14px;font-weight:500;color:#374151;margin-bottom:4px}
.drop-sub{font-size:12px;color:#94A3B8}
.file-selected{background:#EFF6FF;border-color:#3B82F6;border-style:solid}
.file-name{font-size:13px;color:#1D4ED8;font-weight:500;margin-top:8px}
.btn{width:100%;padding:14px;background:#0F172A;color:#fff;border:none;border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;transition:background .2s;display:flex;align-items:center;justify-content:center;gap:8px}
.btn:hover:not(:disabled){background:#1E293B}
.btn:disabled{background:#94A3B8;cursor:not-allowed}
.btn-icon{font-size:16px}
.status{margin-top:16px;padding:12px 16px;border-radius:8px;font-size:13px;display:none}
.status.ok{background:#DCFCE7;color:#15803D;display:block}
.status.err{background:#FEE2E2;color:#991B1B;display:block}
.status.loading{background:#DBEAFE;color:#1D4ED8;display:block}
.steps{margin-top:28px;padding-top:24px;border-top:1px solid #F1F5F9}
.steps-title{font-size:11px;font-weight:600;color:#94A3B8;text-transform:uppercase;letter-spacing:.8px;margin-bottom:12px}
.step{display:flex;gap:12px;margin-bottom:10px;align-items:flex-start}
.step-n{background:#0F172A;color:#fff;border-radius:50%;width:20px;height:20px;font-size:10px;font-weight:600;display:flex;align-items:center;justify-content:center;flex-shrink:0;margin-top:1px}
.step-t{font-size:12px;color:#475569;line-height:1.5}
.progress{display:none;margin-top:16px}
.progress-bar{height:4px;background:#E2E8F0;border-radius:2px;overflow:hidden}
.progress-fill{height:100%;background:#3B82F6;border-radius:2px;width:0%;transition:width .3s}
</style>
</head>
<body>
<div class="card">
  <div class="logo"><span>QUALITÉ & CO LABO</span></div>
  <h1>Générateur Effectifs Post-Vente</h1>
  <p class="subtitle">Importez votre extraction Excel hebdomadaire du labo et téléchargez le fichier SharePoint formaté prêt à déposer.</p>

  <div class="drop-zone" id="dropZone">
    <input type="file" id="fileInput" accept=".xlsx,.xls">
    <div class="drop-icon">📂</div>
    <div class="drop-text">Glissez votre extraction Excel ici</div>
    <div class="drop-sub">ou cliquez pour sélectionner le fichier</div>
    <div class="file-name" id="fileName"></div>
  </div>

  <button class="btn" id="btnGenerer" disabled onclick="generer()">
    <span class="btn-icon">⬇</span> Générer le fichier SharePoint
  </button>

  <div class="progress" id="progressDiv">
    <div class="progress-bar"><div class="progress-fill" id="progressFill"></div></div>
  </div>

  <div class="status" id="status"></div>

  <div class="steps">
    <div class="steps-title">Comment utiliser</div>
    <div class="step"><div class="step-n">1</div><div class="step-t">Importez l'extraction Excel hebdomadaire du labo</div></div>
    <div class="step"><div class="step-n">2</div><div class="step-t">Cliquez sur "Générer" — le fichier se télécharge automatiquement</div></div>
    <div class="step"><div class="step-n">3</div><div class="step-t">Déposez le fichier dans SharePoint — le lien des restaurants ne change pas</div></div>
  </div>
</div>

<script>
const dropZone = document.getElementById('dropZone')
const fileInput = document.getElementById('fileInput')
const fileName  = document.getElementById('fileName')
const btnGen    = document.getElementById('btnGenerer')
const status    = document.getElementById('status')
const progress  = document.getElementById('progressDiv')
const fill      = document.getElementById('progressFill')

fileInput.addEventListener('change', () => {
  if(fileInput.files[0]) selectFile(fileInput.files[0])
})

dropZone.addEventListener('dragover', e => { e.preventDefault(); dropZone.classList.add('over') })
dropZone.addEventListener('dragleave', () => dropZone.classList.remove('over'))
dropZone.addEventListener('drop', e => {
  e.preventDefault(); dropZone.classList.remove('over')
  if(e.dataTransfer.files[0]) selectFile(e.dataTransfer.files[0])
})

function selectFile(file) {
  dropZone.classList.add('file-selected')
  fileName.textContent = '✓ ' + file.name
  btnGen.disabled = false
  status.className = 'status'
  status.textContent = ''
}

function setStatus(msg, type) {
  status.textContent = msg
  status.className = 'status ' + type
}

function animateProgress() {
  fill.style.width = '0%'
  progress.style.display = 'block'
  let w = 0
  const iv = setInterval(() => {
    w = Math.min(w + Math.random() * 8, 90)
    fill.style.width = w + '%'
    if(w >= 90) clearInterval(iv)
  }, 200)
  return iv
}

async function generer() {
  const file = fileInput.files[0]
  if(!file) return

  btnGen.disabled = true
  setStatus('Génération en cours...', 'loading')
  const iv = animateProgress()

  const formData = new FormData()
  formData.append('fichier', file)

  try {
    const resp = await fetch('/generer', { method:'POST', body:formData })
    clearInterval(iv)
    fill.style.width = '100%'

    if(resp.ok) {
      const blob = await resp.blob()
      const cd = resp.headers.get('Content-Disposition') || ''
      const match = cd.match(/filename="?([^"]+)"?/)
      const nomFichier = match ? match[1] : 'Effectifs_SharePoint.xlsx'
      const url = URL.createObjectURL(blob)
      const a = document.createElement('a')
      a.href = url; a.download = nomFichier; a.click()
      URL.revokeObjectURL(url)
      setStatus('✓ Fichier généré avec succès : ' + nomFichier, 'ok')
    } else {
      const err = await resp.json()
      setStatus('Erreur : ' + (err.message || 'Problème lors de la génération'), 'err')
    }
  } catch(e) {
    clearInterval(iv)
    setStatus('Erreur de connexion. Réessayez.', 'err')
  }

  setTimeout(() => { progress.style.display='none'; fill.style.width='0%' }, 1000)
  btnGen.disabled = false
}
</script>
</body>
</html>'''

@app.route('/')
def index():
    return render_template_string(HTML)

@app.route('/generer', methods=['POST'])
def generer():
    if 'fichier' not in request.files:
        return jsonify({'message': 'Aucun fichier reçu'}), 400
    fichier = request.files['fichier']
    if not fichier.filename:
        return jsonify({'message': 'Nom de fichier vide'}), 400
    try:
        output, nom_fichier = generer_excel(fichier)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nom_fichier
        )
    except Exception as e:
        return jsonify({'message': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
